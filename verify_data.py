import sys
import openpyxl
import mysql.connector
from mysql.connector import Error
import re
import json
import csv
from datetime import datetime

def create_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("MySQL Database connection successful")
    except Error as e:
        print(f"The error '{e}' occurred")
    return connection

def sanitize_column_name(name):
    # Replace spaces with underscores, convert to uppercase, and remove special characters
    name = name.replace(" ", "_").upper()
    name = re.sub(r'[^A-Z0-9_]', '', name)
    return name

def read_columns(file_path):
    """Read column names from a file, supporting both CSV and XLSX."""
    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            columns = next(reader)  # Read the first line to get column headers
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path , data_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    else:
        raise ValueError("Unsupported file format")
    return columns

def create_table_from_file(connection, file_path, table_name):
    """Create a table in the database from a file."""
    columns = read_columns(file_path)
    sanitized_columns = [sanitize_column_name(col) for col in columns]
    column_definitions = [f"{col} TEXT" for col in sanitized_columns]


    cursor = connection.cursor()
    try:
        cursor.execute(f"SHOW TABLES LIKE '{table_name}';")
        if cursor.fetchone():
            cursor.execute(f"DROP TABLE {table_name};")
            print(f"Existing table {table_name} dropped.")
    
        create_table_query = f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            {', '.join(column_definitions)}
             );
            """
        cursor.execute(create_table_query)
        print(f"Table {table_name} created successfully")
    except Error as e:
        print(f"The error '{e}' occurred")

def insert_data_from_file(connection, file_path, table_name):
    """Insert data from a file into the database."""
    columns = read_columns(file_path)
    sanitized_columns = [sanitize_column_name(col) for col in columns]
    insert_sql = f"INSERT INTO {table_name} ({', '.join(sanitized_columns)}) VALUES ({', '.join(['%s'] * len(sanitized_columns))})"

    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip the header row
            for row in reader:
                # row = format_date_in_row(row, columns, 'MOIS')
                insert_row(connection, insert_sql, row)
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path ,data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):  # Skip the header row
            cell_values = []
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell_values.append(cell.value.strftime('%m/%d/%y'))  # Format for SQL DATE
                elif isinstance(cell.value, (int, float)):
                    cell_values.append(round(cell.value, 3))
                else:
                    cell_values.append(cell.value)
            # cell_values = format_date_in_row(cell_values, [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))], 'MOIS')
            insert_row(connection, insert_sql, cell_values)


def insert_row(connection, insert_sql, row):
    """Helper function to insert a row into the database."""
    cursor = connection.cursor()
    try:
        formatted_row = [str(cell).strip() if cell is not None else None for cell in row]
        cursor.execute(insert_sql, formatted_row)
        connection.commit()
    except Error as e:
        print(f"Failed to insert row {row}: {e}")

def convert_column_to_date(connection, table_name, column_name):
    """Convert a TEXT column to DATE type after formatting the date."""
    # First, update the column to format dates into 'YYYY-MM-DD'
    format_dates_query = f"""
    UPDATE {table_name}
    SET {column_name} = STR_TO_DATE({column_name}, '%m/%d/%y');
    """
    # Then, modify the column type to DATE
    convert_type_query = f"""
    ALTER TABLE {table_name}
    MODIFY {column_name} DATE;
    """
    cursor = connection.cursor()
    try:
        cursor.execute(format_dates_query)
        connection.commit()
        print(f"Dates in column {column_name} formatted successfully.")
        
        cursor.execute(convert_type_query)
        connection.commit()
        print(f"Column {column_name} converted to DATE successfully.")
    except Error as e:
        print(f"Failed to convert {column_name} to DATE: {e}")

def main():
    if len(sys.argv) != 7:
        print("Usage: python verify_data.py <refFilePath> <annexeFilePath> <host> <db_name> <username> <password>")
        return

    ref_file_path = sys.argv[1]
    annexe_file_path = sys.argv[2]
    host_name = sys.argv[3]
    db_name = sys.argv[4]
    user_name = sys.argv[5]
    user_password = sys.argv[6]

    connection = create_connection(host_name, user_name, user_password, db_name)

    # Create tables for refFile and annexeFile
    create_table_from_file(connection, ref_file_path, "REF_TABLE")
    create_table_from_file(connection, annexe_file_path, "ANNEXE_TABLE")

    # Insert data from Excel files into the tables
    insert_data_from_file(connection, ref_file_path, "REF_TABLE")
    insert_data_from_file(connection, annexe_file_path, "ANNEXE_TABLE")

    convert_column_to_date(connection, "REF_TABLE", "MOIS")
    convert_column_to_date(connection, "ANNEXE_TABLE", "MOIS")

    return(json.dumps({"success": True}))

if __name__ == "__main__":
    main()

