import sys
import csv
import mysql.connector
from mysql.connector import Error
import re
from datetime import datetime
import openpyxl  

def create_connection(host_name, user_name, user_password, db_name):
    """Create a database connection."""
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("MySQL Database connection successful")
    except Error as e:
        print(f"The error '{e}' occurred")
    return connection

def sanitize_column_name(name):
    """Sanitize Excel column names for SQL compatibility."""
    name = name.replace("é", "e").replace("è", "e")
    name = name.replace(" ", "_").upper()
    name = re.sub(r'[^A-Z0_9_]', '', name)
    return name

def read_columns(file_path):
    """Read column names from a file, supporting both CSV and XLSX."""
    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            columns = next(reader)  # Read the first line to get column headers
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path , data_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    else:
        raise ValueError("Unsupported file format")
    return columns

def create_table_from_file(connection, file_path, table_name):
    """Create a table in the database from a file."""
    columns = read_columns(file_path)
    sanitized_columns = [sanitize_column_name(col) for col in columns]
    column_definitions = [f"{col} TEXT" for col in sanitized_columns]

    cursor = connection.cursor()
    try:
        cursor.execute(f"SHOW TABLES LIKE '{table_name}';")
        if cursor.fetchone():
            cursor.execute(f"DROP TABLE {table_name};")
            print(f"Existing table {table_name} dropped.")
    
        create_table_query = f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            {', '.join(column_definitions)}
             );
            """
        cursor.execute(create_table_query)
        print(f"Table {table_name} created successfully")
    except Error as e:
        print(f"The error '{e}' occurred")

def insert_data_from_file(connection, file_path, table_name):
    """Insert data from a file into the database."""
    columns = read_columns(file_path)
    sanitized_columns = [sanitize_column_name(col) for col in columns]
    insert_sql = f"INSERT INTO {table_name} ({', '.join(sanitized_columns)}) VALUES ({', '.join(['%s'] * len(sanitized_columns))})"

    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip the header row
            for row in reader:
                # row = format_date_in_row(row, columns, 'MOIS')
                insert_row(connection, insert_sql, row)
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path ,data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):  # Skip the header row
            cell_values = []
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell_values.append(cell.value.strftime('%m/%d/%y'))  # Format for SQL DATE
                elif isinstance(cell.value, (int, float)):
                    cell_values.append(round(cell.value, 3))
                else:
                    cell_values.append(cell.value)
            # cell_values = format_date_in_row(cell_values, [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))], 'MOIS')
            insert_row(connection, insert_sql, cell_values)


def insert_row(connection, insert_sql, row):
    """Helper function to insert a row into the database."""
    cursor = connection.cursor()
    try:
        formatted_row = [str(cell).strip() if cell is not None else None for cell in row]
        cursor.execute(insert_sql, formatted_row)
        connection.commit()
    except Error as e:
        print(f"Failed to insert row {row}: {e}")

def normalize_analytique_length(connection, table_name):
    """Truncate 'ANALYTIQUE' values to a maximum of 6 characters."""
    update_query = f"""
    UPDATE {table_name}
    SET ANALYTIQUE = LEFT(ANALYTIQUE, 6);
    """
    cursor = connection.cursor()
    try:
        cursor.execute(update_query)
        connection.commit()
        print("ANALYTIQUE column normalized successfully.")
    except Error as e:
        print(f"Failed to normalize ANALYTIQUE column: {e}")

def aggregate_and_modify_data(connection, table_name):
    """Aggregate rows based on 'MOIS' and 'ANALYTIQUE', sum values of 'DEBIT', 'CREDIT', and 'SOLDE', and drop the 'SOC' column."""
    # Aggregate data and update the table
    aggregate_query = f"""
    CREATE TABLE new_table AS
    SELECT MOIS, ANALYTIQUE, SUM(DEBIT) AS DEBIT, SUM(CREDIT) AS CREDIT, SUM(SOLDE) AS SOLDE
    FROM {table_name}
    GROUP BY MOIS, ANALYTIQUE;
    """

    # Drop the original table
    drop_original_table_query = f"DROP TABLE {table_name};"

    # Rename the new table to the original table name
    rename_table_query = f"ALTER TABLE new_table RENAME TO {table_name};"

    cursor = connection.cursor()
    try:
        cursor.execute(aggregate_query)
        cursor.execute(drop_original_table_query)
        cursor.execute(rename_table_query)
        connection.commit()
        print("Data aggregation and modification completed successfully.")
    except Error as e:
        print(f"An error occurred: {e}")

def calculate_and_populate_solde_real(connection, table_name):
    """Calculate the 'SOLDE_REAL' as the difference in 'SOLDE' from one month to the next for each 'ANALYTIQUE'."""
    # Step 1: Add the 'SOLDE_REAL' column to the table if it doesn't exist
    add_column_query = f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS SOLDE_REAL DECIMAL(10,2);"
    cursor = connection.cursor()
    try:
        cursor.execute(add_column_query)
        connection.commit()
        print("'SOLDE_REAL' column added successfully.")
    except Error as e:
        print(f"Failed to add 'SOLDE_REAL' column: {e}")

    # Step 2: Calculate the 'SOLDE_REAL' for each 'MOIS' and 'ANALYTIQUE'
    update_solde_real_query = f"""
    UPDATE {table_name} t1
    SET t1.SOLDE_REAL = t1.SOLDE - COALESCE((
        SELECT t2.SOLDE
        FROM {table_name} t2
        WHERE t2.ANALYTIQUE = t1.ANALYTIQUE 
        AND EXTRACT(YEAR_MONTH FROM STR_TO_DATE(t2.MOIS, '%m/%d/%Y')) = EXTRACT(YEAR_MONTH FROM STR_TO_DATE(t1.MOIS, '%m/%d/%Y')) - 1
    ), 0);
    """
    try:
        cursor.execute(update_solde_real_query)
        connection.commit()
        print("'SOLDE_REAL' updated successfully.")
    except Error as e:
        print(f"Failed to update 'SOLDE_REAL': {e}")

def count_collaborators(file_path):
    """Count the number of rows in the collaborators file, assuming one row per collaborator."""
    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip the header row if there is one
            count = sum(1 for row in reader)
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        count = sum(1 for _ in sheet.iter_rows(min_row=2))  # Skip the header row
    else:
        raise ValueError("Unsupported file format")
    return count

def create_and_populate_solde_part(connection, table_name, collaborators_file):
    """Create the 'SOLDE_PART' column and populate it based on 'SOLDE_REAL' divided by the number of collaborators."""
    # Step 1: Count the number of collaborators
    num_collaborators = count_collaborators(collaborators_file)
    if num_collaborators == 0:
        raise ValueError("No collaborators found in the file.")

    # Step 2: Add the 'SOLDE_PART' column to the table
    add_column_query = f"ALTER TABLE {table_name} ADD COLUMN SOLDE_PART DECIMAL(10,2);"
    cursor = connection.cursor()
    try:
        cursor.execute(add_column_query)
        connection.commit()
        print("'SOLDE_PART' column added successfully.")
    except Error as e:
        print(f"Failed to add 'SOLDE_PART' column: {e}")

    # Step 3: Calculate and update the 'SOLDE_PART' for each row
    update_solde_part_query = f"""
    UPDATE {table_name}
    SET SOLDE_PART = SOLDE_REAL / {num_collaborators};
    """
    try:
        cursor.execute(update_solde_part_query)
        connection.commit()
        print("'SOLDE_PART' updated successfully.")
    except Error as e:
        print(f"Failed to update 'SOLDE_PART': {e}")

def convert_column_to_date(connection, table_name, column_name):
    """Convert a TEXT column to DATE type after formatting the date."""
    # First, update the column to format dates into 'YYYY-MM-DD'
    format_dates_query = f"""
    UPDATE {table_name}
    SET {column_name} = STR_TO_DATE({column_name}, '%m/%d/%y');
    """
    # Then, modify the column type to DATE
    convert_type_query = f"""
    ALTER TABLE {table_name}
    MODIFY {column_name} DATE;
    """
    cursor = connection.cursor()
    try:
        cursor.execute(format_dates_query)
        connection.commit()
        print(f"Dates in column {column_name} formatted successfully.")
        
        cursor.execute(convert_type_query)
        connection.commit()
        print(f"Column {column_name} converted to DATE successfully.")
    except Error as e:
        print(f"Failed to convert {column_name} to DATE: {e}")

def main():
    if len(sys.argv) != 7:
        print("Usage: python preprocess_incharges.py <filePath> <host> <db_name> <username> <password> <collaborators_file>")
        return

    file_path = sys.argv[1]
    host_name = sys.argv[2]
    db_name = sys.argv[3]
    user_name = sys.argv[4]
    user_password = sys.argv[5]
    collaborators_file = sys.argv[6]

    connection = create_connection(host_name, user_name, user_password, db_name)
    table_name = "incharges_table"
    create_table_from_file(connection, file_path, table_name)
    insert_data_from_file(connection, file_path, table_name)
    
    # Convert MOIS column from TEXT to DATE
    convert_column_to_date(connection, table_name, "MOIS")

    normalize_analytique_length(connection, table_name)
    aggregate_and_modify_data(connection, table_name)

    calculate_and_populate_solde_real(connection, table_name)
    create_and_populate_solde_part(connection, table_name, collaborators_file)

if __name__ == "__main__":
    main()

