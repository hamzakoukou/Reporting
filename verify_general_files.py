import sys
import pandas as pd
import openpyxl
import csv
import mysql.connector
from mysql.connector import Error
import re
from datetime import datetime

def create_connection(host_name, user_name, user_password, db_name):
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("MySQL Database connection successful")
        return connection
    except Error as e:
        print(f"Connection Error: {e}")
        return None

def read_columns(file_path):
    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            columns = next(reader)
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    else:
        raise ValueError("Unsupported file format")
    return columns

def sanitize_column_name(name):
    name = name.replace(" ", "_").upper()
    name = re.sub(r'[^A-Z0-9_]', '', name)
    return name

def read_file(file_path):
    if file_path.endswith('.csv'):
        return pd.read_csv(file_path)
    elif file_path.endswith('.xlsx'):
        return pd.read_excel(file_path)
    else:
        raise ValueError("Unsupported file format")

def check_names(connection, table_names):
    cursor = connection.cursor()
    unmatched_names = []

    # Query names from each table and convert them to a set
    names_in_tables = []
    for table_name in table_names:
        cursor.execute(f"SELECT DISTINCT `NOM_PRENOM` FROM `{table_name}`")
        names_in_tables.append(set([name[0].upper() for name in cursor.fetchall()]))

    # Find the intersection of all names across tables
    common_names = set.intersection(*names_in_tables)

    # Identify if any table has names not present in the others
    for idx, names in enumerate(names_in_tables):
        if not names.issubset(common_names):
            unmatched_names.extend(list(names - common_names))

    if unmatched_names:
        print(f"Error: The following names do not match: {unmatched_names}")
        return False, unmatched_names
    return True, None

def delete_tables(connection, table_names):
    cursor = connection.cursor()
    for table_name in table_names:
        cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")
    connection.commit()
    print("Tables deleted due to unmatched names.")

def create_table_from_file(connection, file_path, table_name):
    try:
        # Read columns from the file
        columns = read_columns(file_path)
        print(f"Columns from file: {columns}")

        # Sanitize column names
        sanitized_columns = [sanitize_column_name(col) for col in columns]
        print(f"Sanitized columns: {sanitized_columns}")

        # Create column definitions for SQL
        column_definitions = [f"`{col}` TEXT" for col in sanitized_columns]
        print(f"Column definitions: {column_definitions}")

        # Prepare the cursor
        cursor = connection.cursor()

        # Drop the table if it exists
        drop_table_query = f"DROP TABLE IF EXISTS `{table_name}`;"
        print(f"Executing query: {drop_table_query}")
        cursor.execute(drop_table_query)

        # Create the table
        create_table_query = f"CREATE TABLE `{table_name}` ({', '.join(column_definitions)});"
        print(f"Executing query: {create_table_query}")
        cursor.execute(create_table_query)

        # Commit the transaction
        connection.commit()
        print(f"Table `{table_name}` created successfully")

    except Error as e:
        print(f"Error creating table {table_name}: {e}")

def format_date(value):
    date_formats = ['%m/%d/%y', '%Y-%m-%d', '%m/%d/%Y']  # Add more formats as needed
    for date_format in date_formats:
        try:
            return datetime.strptime(value, date_format).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return value  # Return the value unchanged if it doesn't match any format

def insert_data_from_file(connection, file_path, table_name):
    """Insert data from a file into the database."""
    columns = read_columns(file_path)
    sanitized_columns = [sanitize_column_name(col) for col in columns]
    insert_sql = f"INSERT INTO {table_name} ({', '.join(sanitized_columns)}) VALUES ({', '.join(['%s'] * len(sanitized_columns))})"

    if file_path.endswith('.csv'):
        with open(file_path, newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip the header row
            for row in reader:
                # row = format_date_in_row(row, columns, 'MOIS')
                insert_row(connection, insert_sql, row)
    elif file_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path ,data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):  # Skip the header row
            cell_values = []
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell_values.append(cell.value.strftime('%Y-%m-%d'))  # Format for SQL DATE
                elif isinstance(cell.value, (int, float)):
                    cell_values.append(round(cell.value, 3))
                else:
                    cell_values.append(cell.value)
            # cell_values = format_date_in_row(cell_values, [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))], 'MOIS')
            insert_row(connection, insert_sql, cell_values)

def insert_row(connection, insert_sql, row):
    """Helper function to insert a row into the database."""
    cursor = connection.cursor()
    try:
        formatted_row = [str(cell).strip() if cell is not None else None for cell in row]
        cursor.execute(insert_sql, formatted_row)
        connection.commit()
    except Error as e:
        print(f"Failed to insert row {row}: {e}")

def main():
    print("Starting script...")
    if len(sys.argv) != 9:
        print("Usage: python script.py <host> <db_name> <username> <password> <collaborators_path> <production_path> <charges_path> <ref_table_path>")
        return

    host_name = sys.argv[1]
    db_name = sys.argv[2]
    user_name = sys.argv[3]
    user_password = sys.argv[4]
    collaborators_path = sys.argv[5]
    production_path = sys.argv[6]
    charges_path = sys.argv[7]
    ref_table_path = sys.argv[8]

    print(f"Connecting to database with host: {host_name}, db: {db_name}, user: {user_name}")
    connection = create_connection(host_name, user_name, user_password, db_name)
    if connection is None:
        print("Failed to connect to database.")
        return

    print("Connected to database. Processing files...")

    file_paths = [collaborators_path, production_path, charges_path]
    table_names = ["collaborateurs_table", "production_table", "charges_table"]
    for file_path, table_name in zip(file_paths, table_names):
        print(f"Processing file: {file_path}")
        print(f"Creating table: {table_name}")
        create_table_from_file(connection, file_path, table_name)
        print(f"Inserting data into table: {table_name}")
        insert_data_from_file(connection, file_path, table_name)

    # Check names in the database tables
    check_result, unmatched_names = check_names(connection, table_names)
    if not check_result:
        delete_tables(connection, table_names)
        print(f"Name check failed. Exiting. Unmatched names: {unmatched_names}")
        return

    print("Script completed successfully.")

if __name__ == "__main__":
    main()
